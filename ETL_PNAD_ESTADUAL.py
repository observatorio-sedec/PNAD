
from datetime import datetime
import pandas as pd
import requests as rq 
import pprint
from localidades import estadual
import ssl
import openpyxl
from ajustar_planilha import ajustar_colunas, ajustar_bordas

tabela1209 = 1209
tabela5918 = 5918
tabela6463 = 6463
tabela6482 = 6482

api_populacao = f'https://servicodados.ibge.gov.br/api/v3/agregados/1209/periodos/2022/variaveis/606?{estadual}&classificacao=58[0]'

api_anos = f'https://servicodados.ibge.gov.br/api/v3/agregados/5918/periodos/201201/variaveis/606?localidades=N3[all]&classificacao=58[40288, 114535, 100052, 108875, 99127, 3302]'

api_trab = f'https://servicodados.ibge.gov.br/api/v3/agregados/6463/periodos/201201/variaveis/1641?{estadual}&classificacao=629[32386,32387,32446,32447]'
    
api_potencial = f'https://servicodados.ibge.gov.br/api/v3/agregados/6482/periodos/201201/variaveis/1641?{estadual}&classificacao=604[31751,31752,46254]'

class TLSAdapter(rq.adapters.HTTPAdapter):
    def init_poolmanager(self, *args, **kwargs):
        ctx = ssl.create_default_context()
        ctx.set_ciphers("DEFAULT@SECLEVEL=1")
        ctx.options |= 0x4   # OP_LEGACY_SERVER_CONNECT
        kwargs["ssl_context"] = ctx
        return super(TLSAdapter, self).init_poolmanager(*args, **kwargs)

def requisitando_dados(api):
    with rq.session() as s:
        s.mount("https://", TLSAdapter())
        dados_brutos_api = s.get(api, verify=True)

    # Verificação se a solicitação foi bem-sucedida antes de continuar
    if dados_brutos_api.status_code != 200:
        raise Exception(f"A solicitação à API falhou com o código de status: {dados_brutos_api.status_code}")

    # Verificação se a resposta pode ser convertida para JSON
    try:
        dados_brutos = dados_brutos_api.json()
    except Exception as e:
        raise Exception(f"Erro ao analisar a resposta JSON da API: {str(e)}")
    if len(dados_brutos) < 1:
        dados_brutos = None
        return dados_brutos
    
    return dados_brutos

def extrair_dados(api, tabela_id):
    dados_brutos = requisitando_dados(api)

    if dados_brutos:
        if tabela_id == tabela1209:
            variavel606 = dados_brutos[0]
            return variavel606
        elif tabela_id == tabela5918:
            variavel606 = dados_brutos[0]
            return variavel606
        elif tabela_id == tabela6463:
            variavel1641 = dados_brutos[0]
            return variavel1641
        elif tabela_id == tabela6482:
            variavel1641 = dados_brutos[0]
            return variavel1641
    else:
        return None
    
def tratando_dados1209(variavel1209):
    dados_limpos_1209 = []

    
    id_tabela = variavel1209['id']
    variavel = variavel1209['variavel']
    unidade = variavel1209['unidade']
    dados = variavel1209['resultados']
    
    for ii in dados:
        dados_produto = ii['classificacoes']
        dados_producao = ii['series']

        for iii in dados_produto:
            dados_id_produto = iii['categoria']

            for id_produto, nome_produto in dados_id_produto.items():

                for iv in dados_producao:
                    id = iv['localidade']['id']
                    local = iv['localidade']['nome']
                    dados_ano_producao = iv['serie']
                    

                    for ano, producao in dados_ano_producao.items():
                        producao = producao.replace('-', '0').replace('...', '0')
                        
                        
                        dict = {
                            'id': id,
                            'local': local,
                            #'id_produto': id_produto,
                            'Categoria': nome_produto,
                            variavel: producao,
                            'unidade': unidade,
                            'ano': f'01/01/{ano}'
                            
                        }

                        dados_limpos_1209.append(dict)
    return dados_limpos_1209     
    
def tratando_dados(variavel):
    dados_limpos = []

    # Extraindo informações da variável
    id_tabela = variavel['id']
    variavele = variavel['variavel']
    unidade = variavel['unidade']
    dados = variavel['resultados']

    for item in dados:
        dados_produto = item['classificacoes']
        dados_producao = item['series']
        for classificacao in dados_produto:
            dados_id_produto = classificacao['categoria']

            for id_produto, nome_produto in dados_id_produto.items():
                for serie in dados_producao:
                    id_localidade = serie['localidade']['id']
                    local = serie['localidade']['nome']
                    dados_ano_producao = serie['serie']

                    for ano, producao in dados_ano_producao.items():
                        producao = producao.replace('-', '0').replace('...', '0')

                        partes = ano.split("/")
                        ano_sem_trimestre = int(partes[0][:4])
                        trimestre = int(partes[0][4:6]) 

                        # Criando o dicionário com os dados tratados
                        dados_limpos.append({
                            'id': id_localidade,
                            'local': local,
                            'Categoria': nome_produto,
                            variavele: producao,
                            'unidade': unidade,
                            'ano': f'01/01/{ano_sem_trimestre}',
                            'Trimestre': trimestre,
                            'AnoSedec': f"01/{trimestre * 3 if trimestre * 3 == 12 else f'0{trimestre * 3}'}/{ano_sem_trimestre}"

                        })

    return dados_limpos

def tratando_dados_taxa(variavel):
    dados_limpos = []
    seen = set()  # Usar um set para rastrear entradas únicas
    
    id_tabela = variavel['id']
    variavele = variavel['variavel']
    unidade = variavel['unidade']
    dados = variavel['resultados']

    for item in dados:
        dados_producao = item['series']
        for classificacao in dados_producao:
            dados_id_produto = classificacao['serie']

            for serie in dados_producao:
                id_localidade = serie['localidade']['id']
                local = serie['localidade']['nome']
                dados_ano_producao = serie['serie']

                for ano, producao in dados_ano_producao.items():
                    producao = producao.replace('-', '0').replace('...', '0')

                    partes = ano.split("/")
                    ano_sem_trimestre = int(partes[0][:4])
                    trimestre = int(partes[0][4:6])
                    mes = f"0{trimestre * 3}" if trimestre * 3 < 10 else trimestre * 3
                    
                    entrada_unica = (id_localidade, ano_sem_trimestre, trimestre)
                    if entrada_unica not in seen:
                        seen.add(entrada_unica)  
                        dados_limpos.append({
                            'id': id_localidade,
                            'local': local,
                            variavele: producao,
                            'unidade': unidade,
                            'ano': f'01/01/{ano_sem_trimestre}',
                            'Trimestre': trimestre,
                            'AnoSedec': f'01/{mes}/{ano_sem_trimestre}'
                        })

    return dados_limpos

ano_atual = int(datetime.now().year)

def executando_funcoes(tabela):
    lista_dados1209 = []
    lista_dados5918 = []
    lista_dados6463 = []
    lista_dados6482 = []

    for ano in range(2020, ano_atual+1):  
        if tabela == '1209':
            api_populacao = f'https://servicodados.ibge.gov.br/api/v3/agregados/1209/periodos/{ano}/variaveis/606?localidades=N3[all]&classificacao=58[0]'
            variavel1209 = extrair_dados(api_populacao, tabela1209)
            if variavel1209 is None:
                continue  
            lista_dados1209.extend(tratando_dados1209(variavel1209))
        for tri in range(1, 5):  
            if tabela == '5918':
                api_anos = f'https://servicodados.ibge.gov.br/api/v3/agregados/5918/periodos/{ano}0{tri}/variaveis/606?localidades=N3[all]&classificacao=58[all]'
                variavel5918 = extrair_dados(api_anos, tabela5918)
                if variavel5918 is None:
                    continue
                lista_dados5918.extend(tratando_dados(variavel5918))
                
            elif tabela == '6463':
                api_trab = f'https://servicodados.ibge.gov.br/api/v3/agregados/6463/periodos/{ano}0{tri}/variaveis/1641?localidades=N3[all]&classificacao=629[32386,32387,32446,32447]'
                variavel6463 = extrair_dados(api_trab, tabela6463)
                if variavel6463 is None:
                    continue
                lista_dados6463.extend(tratando_dados(variavel6463))

            elif tabela == '6482':  
                api_potencial = f'https://servicodados.ibge.gov.br/api/v3/agregados/6482/periodos/{ano}0{tri}/variaveis/1641?localidades=N3[all]&classificacao=604[31751,31752,46254]'
                variavel6482 = extrair_dados(api_potencial, tabela6482)
                if variavel6482 is None:
                    continue
                lista_dados6482.extend(tratando_dados(variavel6482))

    if tabela == '1209':
        return lista_dados1209
    elif tabela == '5918':
        return lista_dados5918
    elif tabela == '6463':
        return lista_dados6463
    elif tabela == '6482':
        return lista_dados6482
    else:
        raise ValueError("Tabela desconhecida fornecida")


def executando_estadual():
    lista_dados_8529 = [] 
    for ano in range(2020, ano_atual+1):
        for tri in range(1, 5):
            api_estadual = f'https://servicodados.ibge.gov.br/api/v3/agregados/8529/periodos/{ano}0{tri}/variaveis/12466?localidades=N3[all]'     
            variavel_8529estadual = requisitando_dados(api_estadual)

            if not variavel_8529estadual:
                print(f"A API retornou dados vazios para o ano {ano} e trimestre {tri}.")
                continue  
                
            variavel = variavel_8529estadual[0] 
            novos_dados_8529 = tratando_dados_taxa(variavel)
            lista_dados_8529.extend(novos_dados_8529)

    return lista_dados_8529


def gerando_dataframe(dados_limpos_1209, dados_limpos_5918, dados_limpos_6463, dados_limpos_6482):
    df1209 = pd.DataFrame(dados_limpos_1209)
    df5918 = pd.DataFrame(dados_limpos_5918)
    df6463 = pd.DataFrame(dados_limpos_6463)
    df6482 = pd.DataFrame(dados_limpos_6482)
    
    df5918['População'] = df5918['População'].astype(float)

    
    df5918['População'] = df5918['População'] 
    df6463['Pessoas de 14 anos ou mais de idade'] = df6463['Pessoas de 14 anos ou mais de idade'] 
    df6482['Pessoas de 14 anos ou mais de idade'] = df6482['Pessoas de 14 anos ou mais de idade'] 
    
    df5918[['0 a 13 anos', '14 a 17 anos', '18 a 24 anos', '25 a 39 anos', '40 a 59 anos', '60 anos ou mais']] = None
    linhas_5918_0, linhas_5918_14, linhas_5918_18, linhas_5918_25, linhas_5918_40, linhas_5918_60 = slice(0, 1295), slice(1296, 2591), slice(2592, 3887), slice(3888, 5183), slice(5184, 6479), slice(6480, 7775)
    lista_5918 = [linhas_5918_0, linhas_5918_14, linhas_5918_18, linhas_5918_25, linhas_5918_40, linhas_5918_60]
    lista_mover5918 = ['0 a 13 anos', '14 a 17 anos', '18 a 24 anos', '25 a 39 anos', '40 a 59 anos', '60 anos ou mais']
    for linhas, mover in zip(lista_5918, lista_mover5918):
        df5918.loc[linhas, mover] = df5918.loc[linhas, 'População']
    
    df5918 = df5918.pivot_table(index=['id', 'local', 'ano', 'Trimestre', 'AnoSedec', 'unidade'], columns='Categoria', values='População', aggfunc='first')
    df5918 = df5918.reset_index()
    df5918['Pessoas que podem trabalhar'] = None
    df5918['Pessoas que podem trabalhar'] = df5918['14 a 17 anos'] + df5918['18 a 24 anos'] + df5918['25 a 39 anos'] + df5918['40 a 59 anos'] + df5918['60 anos ou mais']
    
    df6463[['Força de trabalho', 'Ocupado', 'Desocupado', 'Fora da Força de trabalho']] = None
    linhas_6463_F, linhas_6463_O, linhas_6463_DO, linhas_6463_FF = slice(0, 1295), slice(1296, 2591), slice(2592, 3887), slice(3888, 5183)
    df6463.loc[linhas_6463_F, 'Força de trabalho'] = df6463.loc[linhas_6463_F, 'Pessoas de 14 anos ou mais de idade']
    df6463.loc[linhas_6463_O, 'Ocupado'] = df6463.loc[linhas_6463_O, 'Pessoas de 14 anos ou mais de idade']
    df6463.loc[linhas_6463_DO, 'Desocupado'] = df6463.loc[linhas_6463_DO, 'Pessoas de 14 anos ou mais de idade']
    df6463.loc[linhas_6463_FF, 'Fora da Força de trabalho'] = df6463.loc[linhas_6463_FF, 'Pessoas de 14 anos ou mais de idade']
    df6463['Categoria'] = df6463['Categoria'].replace({'Força de trabalho - ocupada': 'Ocupado','Força de trabalho - desocupada': 'Desocupado'})
    
    df6463 = df6463.pivot_table(index=['id', 'local', 'ano', 'Trimestre', 'AnoSedec', 'unidade'], columns='Categoria', values='Pessoas de 14 anos ou mais de idade', aggfunc='first')
    df6463 = df6463.reset_index()

    
    df6482[['Subocupado por insuficiência de horas trabalhadas', 'Força de trabalho potencial', 'Desalentado']] = None
    linhas_6482_H, linhas_6482_FF, linhas_6482_DE =  slice(0, 1295), slice(1296, 2591), slice(2592, 3887)
    df6482['Subocupado por insuficiência de horas trabalhadas'] = df6482.loc[linhas_6482_H, 'Pessoas de 14 anos ou mais de idade']
    df6482['Força de trabalho potencial'] = df6482.loc[linhas_6482_FF, 'Pessoas de 14 anos ou mais de idade']
    df6482['Desalentado'] = df6482.loc[linhas_6482_DE, 'Pessoas de 14 anos ou mais de idade']
    
    df6482 = df6482.pivot_table(index=['id', 'local', 'ano', 'Trimestre', 'AnoSedec', 'unidade'], columns='Categoria', values='Pessoas de 14 anos ou mais de idade', aggfunc='first')
    df6482['Desalentado'] = df6482['Desalentado']
    df6482 = df6482.reset_index()

    return df1209, df5918, df6463, df6482

pp = pprint.PrettyPrinter(indent=4)
dados_limpos_1209 = executando_funcoes('1209')
dados_limpos_5918 = executando_funcoes('5918')
dados_limpos_6463 = executando_funcoes('6463')
dados_limpos_6482 = executando_funcoes('6482')
dataframe1209, dataframe5918, dataframe6463, dataframe6482 = gerando_dataframe(dados_limpos_1209, dados_limpos_5918, dados_limpos_6463, dados_limpos_6482)
dftrab_estadual = pd.merge(dataframe6463, dataframe6482, on=['id', 'local','unidade', 'ano', 'Trimestre', 'AnoSedec'], how='inner')
dftrab_estadual['AnoSedec'] = pd.to_datetime(dftrab_estadual['AnoSedec'], format='%d/%m/%Y')

dados_limpos_8520 = executando_estadual()
df_taxa = pd.DataFrame(dados_limpos_8520)
dataframe5918['AnoSedec'] = pd.to_datetime(dataframe5918['AnoSedec'], format='%d/%m/%Y')


df_taxa.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\taxa.xlsx', index=False)
dataframe1209.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\População ESTADUAL.xlsx', index=False)
dataframe5918.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\Idade ESTADUAL.xlsx', index=False)
dftrab_estadual.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\Trabalho ESTADUAL.xlsx', index=False)


planilha_principal = openpyxl.Workbook()

wb_1209 = openpyxl.load_workbook('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\População ESTADUAL.xlsx')
wb_5918 = openpyxl.load_workbook('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\Idade ESTADUAL.xlsx')
wb_trab = openpyxl.load_workbook('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\Trabalho ESTADUAL.xlsx')

aba_1209 = planilha_principal.create_sheet("População Total")
aba_5918 = planilha_principal.create_sheet("Pessoas aptam a trabalhar")
aba_trab = planilha_principal.create_sheet("Trabalho")

for linha in wb_1209.active.iter_rows(values_only=True):
    aba_1209.append(linha)

for linha in wb_5918.active.iter_rows(values_only=True):
    aba_5918.append(linha)
    
for linha in wb_trab.active.iter_rows(values_only=True):
    aba_trab.append(linha)
    
for aba in planilha_principal.sheetnames:
    if aba not in ["População Total", "Pessoas aptam a trabalhar", "Trabalho"]:
        del planilha_principal[aba]
        
ajustar_bordas(planilha_principal)

lista_aba = [aba_1209, aba_5918, aba_trab]
for abas in lista_aba:
    ajustar_colunas(abas)
    
planilha_principal.save("C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas Tratadas\\PNAD ESTADUAL.xlsx")   

if __name__ == '__main__':
    from sql import executar_sql 
    executar_sql()
