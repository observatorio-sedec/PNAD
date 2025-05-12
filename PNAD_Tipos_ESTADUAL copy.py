import pandas as pd
import requests as rq 
import pprint
import sqlite3
from localidades import estadual
import ssl
from Google import Create_Service
from googleapiclient.http import MediaFileUpload
import openpyxl
from ajustar_planilha import ajustar_colunas, ajustar_bordas

tabela5917 = 5917
tabela4093 = 4093
tabela6398 = 6398
tabela5918 = 5918
tabela4094 = 4094
tabela6388 = 6388
tabela6399 = 6399
tabela6403 = 6403
tabela6402 = 6402
tabela5919 = 5919
tabela4095 = 4095
tabelaespecial = 40952

api_populacao_sexo = f'https://servicodados.ibge.gov.br/api/v3/agregados/5917/periodos/201301|201302|201303|201304|201401|201402|201403|201404|201501|201502|201503|201504|201601|201602|201603|201604|201701|201702|201703|201704|201801|201802|201803|201804|201901|201902|201903|201904|202001|202202|202203|202204|202301|202302|202303/variaveis/606?{estadual}&classificacao=2[4,5]'
api_trab_sexo = f'https://servicodados.ibge.gov.br/api/v3/agregados/4093/periodos/201301|201302|201303|201304|201401|201402|201403|201404|201501|201502|201503|201504|201601|201602|201603|201604|201701|201702|201703|201704|201801|201802|201803|201804|201901|201902|201903|201904|202001|202002|202003|202004|202101|202102|202103|202104|202201|202202|202203|202204|202301|202302|202303|202304/variaveis/1641|4088|4090|4092|4094?localidades=N3[all]&classificacao=2[4,5]'
api_forca_sexo = f'https://servicodados.ibge.gov.br/api/v3/agregados/6398/periodos/201301|201302|201303|201304|201401|201402|201403|201404|201501|201502|201503|201504|201601|201602|201603|201604|201701|201702|201703|201704|201801|201802|201803|201804|201901|201902|201903|201904|202001|202002|202003|202004|202101|202102|202103|202104|202201|202202|202203|202204|202301|202302|202303|202304/variaveis/8344|8346?localidades=N3[all]&classificacao=2[4,5]'

api_populacao_idade= f'https://servicodados.ibge.gov.br/api/v3/agregados/5918/periodos/201201|201202|201203|201204|201301|201302|201303|201304|201401|201402|201403|201404|201501|201502|201503|201504|201601|201602|201603|201604|201701|201702|201703|201704|201801|201802|201803|201804|201901|201902|201903|201904|202001|202002|202003|202004|202101|202102|202103|202104|202201|202202|202203|202204|202301|202302|202303|202304/variaveis/606?localidades=N3[all]&classificacao=58[95253,114535,100052,108875,99127,3302]'
api_anos_idade = f' https://servicodados.ibge.gov.br/api/v3/agregados/{tabela4094}/periodos/201201|201202|201203|201204|201301|201302|201303|201304|201401|201402|201403|201404|201501|201502|201503|201504|201601|201602|201603|201604|201701|201702|201703|201704|201801|201802|201803|201804|201901|201902|201903|201904|202001|202002|202003|202004|202101|202102|202103|202104|202201|202202|202203|202204|202301|202302|202303|202304/variaveis/1641|4088|4090|4092|4094?{estadual}&classificacao=58[114535,100052,108875,99127,3302]'
api_trab_idade = f'https://servicodados.ibge.gov.br/api/v3/agregados/{tabela6399}/periodos/201201|201202|201203|201204|201301|201302|201303|201304|201401|201402|201403|201404|201501|201502|201503|201504|201601|201602|201603|201604|201701|201702|201703|201704|201801|201802|201803|201804|201901|201902|201903|201904|202001|202202|202203|202204|202301|202302|202303|202304/variaveis/8344|8346?{estadual}&classificacao=58[114535,100052,108875,99127,3302]'

api_populacao_raca= f'https://servicodados.ibge.gov.br/api/v3/agregados/{tabela6403}/periodos/201201|201202|201203|201204|201301|201302|201303|201304|201401|201402|201403|201404|201501|201502|201503|201504|201601|201602|201603|201604|201701|201702|201703|201704|201801|201802|201803|201804|201901|201902|201903|201904|202001|202002|202003|202004|202101|202102|202103|202104|202201|202202|202203|202204|202301|202302|202303|202304/variaveis/606?{estadual}&classificacao=86[2776,2777,2779]'
api_trab_raca = f'https://servicodados.ibge.gov.br/api/v3/agregados/6402/periodos/201201|201202|201203|201204|201301|201302|201303|201304|201401|201402|201403|201404|201501|201502|201503|201504|201601|201602|201603|201604|201701|201702|201703|201704|201801|201802|201803|201804|201901|201902|201903|201904|202001|202002|202003|202004|202101|202102|202103|202104|202201|202202|202203|202204|202301|202302|202303|202304/variaveis/1641|4088|4090|4092|4094?localidades=N3[all]&classificacao=86[2776,2777,2779]'

api_populacao_alfab= f'https://servicodados.ibge.gov.br/api/v3/agregados/{tabela5919}/periodos/201201|201202|201203|201204|201301|201302|201303|201304|201401|201402|201403|201404|201501|201502|201503|201504|201601|201602|201603|201604|201701|201702|201703|201704|201801|201802|201803|201804|201901|201902|201903|201904|202001|202002|202003|202004|202101|202102|202103|202104|202201|202202|202203|202204|202301|202302|202303|202304/variaveis/606?{estadual}&classificacao=1568[120706,11779,11628,11629,11630,11631,11632,11626]'
api_trab_alfab = f'https://servicodados.ibge.gov.br/api/v3/agregados/4095/periodos/201201|201202|201203|201204|201301|201302|201303|201304|201401|201402|201403|201404|201501|201502|201503|201504|201601|201602|201603|201604|201701|201702|201703|201704|201801|201802|201803|201804|201901|201902|201903|201904|202001|202002|202003|202004|202101|202102|202103|202104|202201|202202|202203|202204|202301|202302|202303|202304/variaveis/1641|4088|4090|4092?{estadual}&classificacao=1568[120706,11779,11628,11629,11630,11631,11632,11626]'
api_trab2_alfab = 'https://servicodados.ibge.gov.br/api/v3/agregados/4095/periodos/201201|201202|201203|201204|201301|201302|201303|201304|201401|201402|201403|201404|201501|201502|201503|201504|201601|201602|201603|201604|201701|201702|201703|201704|201801|201802|201803|201804|201901|201902|201903|201904|202001|202002|202003|202004|202101|202102|202103|202104|202201|202202|202203|202204|202301|202302|202303|202304/variaveis/4094?localidades=N3[all]&classificacao=1568[120706,11779,11628,11629,11630,11631,11632,11626]'


class TLSAdapter(rq.adapters.HTTPAdapter):
    def init_poolmanager(self, *args, **kwargs):
        ctx = ssl.create_default_context()
        ctx.set_ciphers("DEFAULT@SECLEVEL=1")
        ctx.options |= 0x4   # OP_LEGACY_SERVER_CONNECT
        kwargs["ssl_context"] = ctx
        return super(TLSAdapter, self).init_poolmanager(*args, **kwargs)

def requisitando_dados(api):
    with rq.session() as s:
        s.mount("https://", TLSAdapter())
        dados_brutos_api = s.get(api, verify=True)

    # Verificação se a solicitação foi bem-sucedida antes de continuar
    if dados_brutos_api.status_code != 200:
        raise Exception(f"A solicitação à API falhou com o código de status: {dados_brutos_api.status_code}")

    # Verificação se a resposta pode ser convertida para JSON
    try:
        dados_brutos = dados_brutos_api.json()
    except Exception as e:
        raise Exception(f"Erro ao analisar a resposta JSON da API: {str(e)}")

    return dados_brutos

def extrair_dados(api, tabela_id):
    dados_brutos = requisitando_dados(api)
    lista_tabela = [tabela5917, tabela5918, tabela6403, tabela5919]
    if dados_brutos:
        for tabelas in lista_tabela:
            if tabela_id == tabelas:
                variavel = dados_brutos[0]
                return variavel
        if tabela_id == 4093 or tabela_id == 4094 or tabela_id == 6402:
            variavel1641 = dados_brutos[0]
            variavel4088 = dados_brutos[1]
            variavel4090 = dados_brutos[2]
            variavel4092 = dados_brutos[3]
            variavel4094 = dados_brutos[4]
            return variavel1641, variavel4088, variavel4090, variavel4092, variavel4094
        elif tabela_id == 4095:
            variavel1641 = dados_brutos[0]
            variavel4088 = dados_brutos[1]
            variavel4090 = dados_brutos[2]
            variavel4092 = dados_brutos[3]
            return variavel1641, variavel4088, variavel4090, variavel4092
        elif tabela_id == 6398 or tabela_id == 6399:
            variavel8344 = dados_brutos[0]
            variavel8346 = dados_brutos[1]
            return variavel8344, variavel8346
        elif tabela_id == 40952:
            variavel4094 = dados_brutos[0]
            return variavel4094
    else:
        pass
        
def tratando_dados(variavel):
    dados_limpos = []

    
    id_tabela = variavel['id']
    variavele = variavel['variavel']
    unidade = variavel['unidade']
    dados = variavel['resultados']
    
    for ii in dados:
        dados_produto = ii['classificacoes']
        dados_producao = ii['series']

        for iii in dados_produto:
            dados_id_produto = iii['categoria']

            for id_produto, nome_produto in dados_id_produto.items():

                for iv in dados_producao:
                    id = iv['localidade']['id']
                    local = iv['localidade']['nome']
                    dados_ano_producao = iv['serie']

                    for ano, producao in dados_ano_producao.items():
                        producao = producao.replace('-', '0').replace('...', '0')
                        
                        partes = ano.split("/")
                        ano_sem_trimestre = int(partes[0][:4])
                        trimestre = int(partes[0][4:6])
                        
                        dict = {
                            'id': id,
                            'local': local,
                            #'id_produto': id_produto,
                            'Categoria': nome_produto,
                            variavele: producao,
                            'unidade': unidade,
                            'ano': f'01/01/{ano_sem_trimestre}',
                            'Trimestre': trimestre,
                            'AnoSedec': f'01/{trimestre * 3}/{ano_sem_trimestre}'
                        }

                        dados_limpos.append(dict)
    return dados_limpos 

def tratando_dados_cinco(variavel1641, variavel4088, variavel4090, variavel4092, variavel4094):
    dados_limpos1641 = []
    dados_limpos4088 = []
    dados_limpos4090 = []
    dados_limpos4092  = []
    dados_limpos4094 = []
    
    variaveis = [variavel1641, variavel4088, variavel4090, variavel4092, variavel4094]
    for i in variaveis:
        id_tabela = i['id']
        variavel = i['variavel']
        unidade = i['unidade']
        dados = i['resultados']
        
        for ii in dados:
            dados_produto = ii['classificacoes']
            dados_producao = ii['series']

            for iii in dados_produto:
                dados_id_produto = iii['categoria']

                for id_produto, nome_produto in dados_id_produto.items():

                    for iv in dados_producao:
                        id = iv['localidade']['id']
                        local = iv['localidade']['nome']
                        dados_ano_producao = iv['serie']
                        

                        for ano, producao in dados_ano_producao.items():
                            producao = producao.replace('-', '0').replace('...', '0')
                            
                            partes = ano.split("/")
                            ano_sem_trimestre = int(partes[0][:4])
                            trimestre = int(partes[0][4:6])
                            
                            dict = {
                                'id': id,
                                'local': local,
                              #  'id_produto': id_produto,
                                'Categoria': nome_produto,
                                variavel: producao,
                                'unidade': unidade,
                                'ano': f'01/01/{ano_sem_trimestre}',
                                'Trimestre': trimestre,
                                'AnoSedec': f'01/{trimestre * 3}/{ano_sem_trimestre}'
                            }
                            if id_tabela == '1641':
                                dados_limpos1641.append(dict)
                            elif id_tabela == '4088':
                                dados_limpos4088.append(dict)
                            elif id_tabela == '4090':
                                dados_limpos4090.append(dict)
                            elif id_tabela == '4092':
                                dados_limpos4092.append(dict)
                            elif id_tabela == '4094':
                                dados_limpos4094.append(dict)
                                
    return dados_limpos1641,  dados_limpos4088, dados_limpos4090, dados_limpos4092, dados_limpos4094

def tratando_dados_quatro(variavel4088, variavel4090, variavel4092, variavel4094):
    dados_limpos4088 = []
    dados_limpos4090 = []
    dados_limpos4092  = []
    dados_limpos4094 = []
    
    variaveis = [variavel4088, variavel4090, variavel4092, variavel4094]
    for i in variaveis:
        id_tabela = i['id']
        variavele = i['variavel']
        unidade = i['unidade']
        dados = i['resultados']
        
        for ii in dados:
            dados_produto = ii['classificacoes']
            dados_producao = ii['series']

            for iii in dados_produto:
                dados_id_produto = iii['categoria']

                for id_produto, nome_produto in dados_id_produto.items():

                    for iv in dados_producao:
                        id = iv['localidade']['id']
                        local = iv['localidade']['nome']
                        dados_ano_producao = iv['serie']
                        

                        for ano, producao in dados_ano_producao.items():
                            producao = producao.replace('-', '0').replace('...', '0')
                            
                            partes = ano.split("/")
                            ano_sem_trimestre = int(partes[0][:4])
                            trimestre = int(partes[0][5:6])
                            
                            dict = {
                                'id': id,
                                'local': local,
                                #'id_produto': id_produto,
                                'Categoria': nome_produto,
                                variavele: producao,
                                'unidade': unidade,
                                'ano': f'01/01/{ano_sem_trimestre}',
                                'Trimestre': trimestre,
                                'AnoSedec': f'01/{trimestre * 3}/{ano_sem_trimestre}'
                            }
                            
                            if id_tabela == '4088':
                                dados_limpos4088.append(dict)
                            elif id_tabela == '4090':
                                dados_limpos4090.append(dict)
                            elif id_tabela == '4092':
                                dados_limpos4092.append(dict)
                            elif id_tabela == '4094':
                                dados_limpos4094.append(dict)
                                
    return dados_limpos4088, dados_limpos4090, dados_limpos4092, dados_limpos4094

def tratando_dados_especial(variavel1641, variavel4088, variavel4090, variavel4092):
    dados_limpos1641 = []
    dados_limpos4088 = []
    dados_limpos4090 = []
    dados_limpos4092  = []
    
    
    variaveis = [variavel1641, variavel4088, variavel4090, variavel4092]
    for i in variaveis:
        id_tabela = i['id']
        variavele = i['variavel']
        unidade = i['unidade']
        dados = i['resultados']
        
        for ii in dados:
            dados_produto = ii['classificacoes']
            dados_producao = ii['series']

            for iii in dados_produto:
                dados_id_produto = iii['categoria']

                for id_produto, nome_produto in dados_id_produto.items():

                    for iv in dados_producao:
                        id = iv['localidade']['id']
                        local = iv['localidade']['nome']
                        dados_ano_producao = iv['serie']
                        

                        for ano, producao in dados_ano_producao.items():
                            producao = producao.replace('-', '0').replace('...', '0')
                            
                            partes = ano.split("/")
                            ano_sem_trimestre = int(partes[0][:4])
                            trimestre = int(partes[0][4:6])
                            
                            dict = {
                                'id': id,
                                'local': local,
                                #'id_produto': id_produto,
                                'Categoria': nome_produto,
                                variavele: producao,
                                'unidade': unidade,
                                'ano': f'01/01/{ano_sem_trimestre}',
                                'Trimestre': trimestre,
                                'AnoSedec': f'01/{trimestre * 3}/{ano_sem_trimestre}'
                            }
                            
                            if id_tabela == '1641':
                                dados_limpos1641.append(dict)
                            elif id_tabela == '4088':
                                dados_limpos4088.append(dict)
                            elif id_tabela == '4090':
                                dados_limpos4090.append(dict)
                            elif id_tabela == '4092':
                                dados_limpos4092.append(dict)
                                
    return  dados_limpos1641, dados_limpos4088, dados_limpos4090, dados_limpos4092

def tratando_dados_dois(variavel8344, variave8346):
    dados_limpos8344  = []
    dados_limpos8346 = []
    
    variaveis = [variavel8344, variave8346]
    for i in variaveis:
        id_tabela = i['id']
        variavel = i['variavel']
        unidade = i['unidade']
        dados = i['resultados']
        
        for ii in dados:
            dados_produto = ii['classificacoes']
            dados_producao = ii['series']

            for iii in dados_produto:
                dados_id_produto = iii['categoria']

                for id_produto, nome_produto in dados_id_produto.items():

                    for iv in dados_producao:
                        id = iv['localidade']['id']
                        local = iv['localidade']['nome']
                        dados_ano_producao = iv['serie']
                        

                        for ano, producao in dados_ano_producao.items():
                            producao = producao.replace('-', '0').replace('...', '0')
                            
                            partes = ano.split("/")
                            ano_sem_trimestre = int(partes[0][:4])
                            trimestre = int(partes[0][4:6])
                            
                            
                            dict = {
                                'id': id,
                                'local': local,
                                # 'id_produto': id_produto,
                                'Categoria': nome_produto,
                                variavel: producao,
                                'unidade': unidade,
                                'ano': f'01/01/{ano_sem_trimestre}',
                                'Trimestre': trimestre,
                                'AnoSedec': f'01/{trimestre * 3}/{ano_sem_trimestre}'
                            }
                            
                            if id_tabela == '8344':
                                dados_limpos8344.append(dict)
                            elif id_tabela == '8346':
                                dados_limpos8346.append(dict)

    return dados_limpos8344, dados_limpos8346

def executando_funcoes():
    variavel_pop_sexo = extrair_dados(api_populacao_sexo, tabela5917)
    variavel_trab_sexo_1641, variavel_trab_sexo_4088, variavel_trab_sexo_4090, variavel_trab_sexo_4092, variavel_trab_sexo_4094 = extrair_dados(api_trab_sexo, tabela4093)
    variavel_forca_sexo_8344, variavel_forca_sexo_8346 = extrair_dados(api_forca_sexo, tabela6398)
    
    variavel_pop_idade = extrair_dados(api_populacao_idade, tabela5918)
    variavel_anos_idade_1641, variavel_anos_idade_4088, variavel_anos_idade_4090, variavel_anos_idade_4092, variavel_anos_sexo_4094 = extrair_dados(api_anos_idade, tabela4094)
    variavel_trab_idade_8344,  variavel_trab_idade_8346 = extrair_dados(api_trab_idade, tabela6399)
    
    variavel_pop_raca = extrair_dados(api_populacao_raca, tabela6403)
    variavel_trab_raca_1641, variavel_trab_raca_4088, variavel_trab_raca_4090, variavel_trab_raca_4092, variavel_trab_raca_4094 = extrair_dados(api_trab_raca, tabela6402)
    
    variavel_pop_alfab = extrair_dados(api_populacao_alfab, tabela5919)
    variavel_trab_alfab_1641, variavel_trab_alfab_4088, variavel_trab_alfab_4090, variavel_trab_alfab_4092 = extrair_dados(api_trab_alfab, tabela4095)
    variavel_trab_alfab_4094 = extrair_dados(api_trab2_alfab, tabelaespecial)
    
    dados_limpos_pop_sexo = tratando_dados(variavel_pop_sexo)
    dados_limpos1641_trab_sexo,  dados_limpos4088_trab_sexo, dados_limpos4090_trab_sexo, dados_limpos4092_trab_sexo, dados_limpos4094_trab_sexo = tratando_dados_cinco(variavel_trab_sexo_1641, variavel_trab_sexo_4088, variavel_trab_sexo_4090, variavel_trab_sexo_4092, variavel_trab_sexo_4094)
    dados_limpos8344_forca_sexo, dados_limpos8346_forca_sexo = tratando_dados_dois(variavel_forca_sexo_8344, variavel_forca_sexo_8346)
    
    dados_limpos_pop_idade = tratando_dados(variavel_pop_idade)
    dados_limpos1641_anos_idade,  dados_limpos4088_anos_idade, dados_limpos4090_anos_idade, dados_limpos4092_anos_idade, dados_limpos4094_anos_idade = tratando_dados_cinco(variavel_anos_idade_1641, variavel_anos_idade_4088, variavel_anos_idade_4090, variavel_anos_idade_4092, variavel_anos_sexo_4094)
    dados_limpos8344_forca_idade, dados_limpos8346_trab_idade = tratando_dados_dois(variavel_trab_idade_8344,  variavel_trab_idade_8346)
    
    dados_limpos_pop_raca = tratando_dados(variavel_pop_raca)
    dados_limpos1641_trab_raca, dados_limpos4088_trab_raca, dados_limpos4090_trab_raca, dados_limpos4092_trab_raca, dados_limpos4094_trab_raca = tratando_dados_cinco(variavel_trab_raca_1641, variavel_trab_raca_4088, variavel_trab_raca_4090, variavel_trab_raca_4092, variavel_trab_raca_4094)
    
    dados_limpos_pop_alfab = tratando_dados(variavel_pop_alfab)
    dados_limpos1641_trab_alfab,  dados_limpos4088_trab_alfab, dados_limpos4090_trab_alfab, dados_limpos4092_trab_alfab = tratando_dados_especial(variavel_trab_alfab_1641, variavel_trab_alfab_4088, variavel_trab_alfab_4090, variavel_trab_alfab_4092)
    dados_limpos4094_trab_alfab = tratando_dados(variavel_trab_alfab_4094)
    
    return dados_limpos_pop_sexo, dados_limpos1641_trab_sexo,  dados_limpos4088_trab_sexo, dados_limpos4090_trab_sexo, dados_limpos4092_trab_sexo, dados_limpos4094_trab_sexo, \
        dados_limpos8344_forca_sexo, dados_limpos8346_forca_sexo, dados_limpos_pop_idade, dados_limpos1641_anos_idade,  dados_limpos4088_anos_idade, dados_limpos4090_anos_idade, dados_limpos4092_anos_idade, dados_limpos4094_anos_idade, \
            dados_limpos8344_forca_idade, dados_limpos8346_trab_idade, dados_limpos_pop_raca, dados_limpos1641_trab_raca, dados_limpos4088_trab_raca, dados_limpos4090_trab_raca, dados_limpos4092_trab_raca, dados_limpos4094_trab_raca, \
                dados_limpos_pop_alfab, dados_limpos1641_trab_alfab, dados_limpos4088_trab_alfab, dados_limpos4090_trab_alfab, dados_limpos4092_trab_alfab, dados_limpos4094_trab_alfab
                
def gerando_dataframePOP(dados_limpos_pop_sexo, dados_limpos_pop_idade, dados_limpos_pop_raca, dados_limpos_pop_alfab):
    dfpop_sexo = pd.DataFrame(dados_limpos_pop_sexo)
    dfpop_idade = pd.DataFrame(dados_limpos_pop_idade)
    dfpop_raca = pd.DataFrame(dados_limpos_pop_raca)
    dfpop_alfab = pd.DataFrame(dados_limpos_pop_alfab)
    
    dfpop_idade['População Geral'] = None
    linhapop = slice(0, 1296)
    linhainfo = slice(0 , 7776)
    valor_populacao = dfpop_idade.iloc[linhapop]['População'].tolist()

    for index_info, row_info in dfpop_idade.iloc[linhainfo].iterrows():
        dfpop_idade.at[index_info, 'População Geral'] = valor_populacao[index_info % len(valor_populacao)]
    return dfpop_sexo, dfpop_idade, dfpop_raca, dfpop_alfab

def gerando_dataframe_cinco(dados_limpos1641,  dados_limpos4088, dados_limpos4090, dados_limpos4092, dados_limpos4094):
    
    df1641 = pd.DataFrame(dados_limpos1641)
    df4088 = pd.DataFrame(dados_limpos4088)
    df4090 = pd.DataFrame(dados_limpos4090)
    df4092 = pd.DataFrame(dados_limpos4092)
    df4094 = pd.DataFrame(dados_limpos4094)
    
    df = pd.merge(df1641, df4088, on=['id', 'local', 'Categoria','unidade', 'ano', 'Trimestre', 'AnoSedec'], how='inner')
    df = pd.merge(df, df4090, on=['id', 'local', 'Categoria', 'unidade', 'ano', 'Trimestre', 'AnoSedec'], how='inner')
    df = pd.merge(df, df4092, on=['id', 'local', 'Categoria','unidade', 'ano', 'Trimestre', 'AnoSedec'], how='inner')
    df = pd.merge(df, df4094, on=['id', 'local', 'Categoria', 'unidade', 'ano', 'Trimestre', 'AnoSedec'], how='inner')
    df.columns = df.columns.str.strip()
    df.rename(columns={
        'Pessoas de 14 anos ou mais de idade, na força de trabalho, na semana de referência': 'Força de Trabalho',
        'Pessoas de 14 anos ou mais de idade ocupadas na semana de referência': 'Ocupadas',
        'Pessoas de 14 anos ou mais de idade, desocupadas na semana de referência': 'Desocupadas',
        'Pessoas de 14 anos ou mais de idade, fora da força de trabalho, na semana de referência': 'Fora da Força de Trabalho'
    }, inplace=True)

    
    return df

def gerando_dataframe_tres(dados_limpos8344, dados_limpos8346):
    df8344 = pd.DataFrame(dados_limpos8344)
    df8346 = pd.DataFrame(dados_limpos8346)
    
    df = pd.merge(df8344, df8346, on=['id', 'local', 'Categoria','unidade', 'ano', 'Trimestre', 'AnoSedec'], how='inner')
    df.rename(columns={
        'Pessoas de 14 anos ou mais de idade, subocupadas por insuficiência de horas trabalhadas': 'Subocupadas por insuficiência de horas trabalhadas',
        'Pessoas de 14 anos ou mais de idade, na força de trabalho potencial':'Força de trabalho potencial'
        }, inplace=True)
    
    # colunas_convertes = ['Subocupadas por insuficiência de horas trabalhadas', 'Força de trabalho potencial']
    # for coluna in colunas_convertes:
    #     df[coluna] = df[coluna].astype(int)

    return df

def gerando_dataframe_quatro(dados_limpos4088, dados_limpos4090, dados_limpos4092, dados_limpos4094):
    df4088 = pd.DataFrame(dados_limpos4088)
    df4090 = pd.DataFrame(dados_limpos4090)
    df4092 = pd.DataFrame(dados_limpos4092)
    df4094 = pd.DataFrame(dados_limpos4094)
    df = pd.merge(df4088, df4090, on=['id', 'local', 'Categoria', 'unidade', 'ano', 'Trimestre', 'AnoSedec'], how='inner')
    df = pd.merge(df, df4092, on=['id', 'local', 'Categoria','unidade', 'ano', 'Trimestre', 'AnoSedec'], how='inner')
    df = pd.merge(df, df4094, on=['id', 'local', 'Categoria', 'unidade', 'ano', 'Trimestre', 'AnoSedec'], how='inner')
    df.rename(columns={
        'Pessoas de 14 anos ou mais de idade, na força de trabalho, na semana de referência': 'Força de Trabalho',
        'Pessoas de 14 anos ou mais de idade ocupadas na semana de referência': 'Ocupadas',
        'Pessoas de 14 anos ou mais de idade, desocupadas na semana de referência': 'Desocupadas',
        'Pessoas de 14 anos ou mais de idade, fora da força de trabalho, na semana de referência': 'Fora da Força de Trabalho'
    }, inplace=True)
    # colunas_convertes = ['Pessoas de 14 anos ou mais de idade', 'Ocupadas', 'Desocupadas', 'Força de Trabalho']
    # for coluna in colunas_convertes:
    #     df[coluna] = df[coluna].astype(int)
    # return df
    
dados_limpos_pop_sexo, dados_limpos1641_trab_sexo,  dados_limpos4088_trab_sexo, dados_limpos4090_trab_sexo, dados_limpos4092_trab_sexo, dados_limpos4094_trab_sexo, \
    dados_limpos8344_forca_sexo, dados_limpos8346_forca_sexo, dados_limpos_pop_idade, dados_limpos1641_anos_idade, dados_limpos4088_anos_idade, dados_limpos4090_anos_idade, dados_limpos4092_anos_idade, dados_limpos4094_anos_idade, \
         dados_limpos8344_forca_idade, dados_limpos8346_trab_idade, dados_limpos_pop_raca, dados_limpos1641_trab_raca, dados_limpos4088_trab_raca, dados_limpos4090_trab_raca, dados_limpos4092_trab_raca, dados_limpos4094_trab_raca,\
             dados_limpos_pop_alfab, dados_limpos1641_trab_alfab, dados_limpos4088_trab_alfab, dados_limpos4090_trab_alfab, dados_limpos4092_trab_alfab, dados_limpos4094_trab_alfab  = executando_funcoes()
             
dfpop_sexo, dfpop_idade, dfpop_raca, dfpop_alfab = gerando_dataframePOP(dados_limpos_pop_sexo, dados_limpos_pop_idade, dados_limpos_pop_raca, dados_limpos_pop_alfab)
dftrab_sexo  = gerando_dataframe_cinco(dados_limpos1641_trab_sexo,  dados_limpos4088_trab_sexo, dados_limpos4090_trab_sexo, dados_limpos4092_trab_sexo, dados_limpos4094_trab_sexo)
dfforca_sexo = gerando_dataframe_tres(dados_limpos8344_forca_sexo, dados_limpos8346_forca_sexo)

dfanos_idade = gerando_dataframe_cinco(dados_limpos1641_anos_idade,  dados_limpos4088_anos_idade, dados_limpos4090_anos_idade, dados_limpos4092_anos_idade, dados_limpos4094_anos_idade)

dfforca_idade = gerando_dataframe_tres(dados_limpos8344_forca_idade, dados_limpos8346_trab_idade)
dftrab_raca = gerando_dataframe_cinco(dados_limpos1641_trab_raca, dados_limpos4088_trab_raca, dados_limpos4090_trab_raca, dados_limpos4092_trab_raca, dados_limpos4094_trab_raca)
dftrab_alfab = gerando_dataframe_cinco(dados_limpos1641_trab_alfab, dados_limpos4088_trab_alfab, dados_limpos4090_trab_alfab, dados_limpos4092_trab_alfab, dados_limpos4094_trab_alfab)


dfpop_sexo.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\População SEXO estadual.xlsx', index=False)
dftrab_sexo.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\trabalho SEXO estadual.xlsx', index=False)
dfforca_sexo.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\Força trabalho SEXO estadual.xlsx', index=False)


#JUNTANDO DATAFRAMES
df_sexo = pd.merge(dfpop_sexo, dftrab_sexo, on=['id', 'local', 'Categoria', 'unidade', 'ano', 'Trimestre', 'AnoSedec'], how='inner')
df_sexo = pd.merge(df_sexo, dfforca_sexo, on=['id', 'local', 'Categoria', 'unidade', 'ano', 'Trimestre', 'AnoSedec'], how='inner')
df_sexo.to_excel("C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas Tratadas\\Tipo SEXO estadual.xlsx")


dfpop_idade.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\Populaçao IDADE estadual.xlsx', index=False)
dfanos_idade.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\Anos IDADE estadual.xlsx', index=False)


dfforca_idade.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\Força trabalho IDADE estadual.xlsx', index=False)
df_idade = pd.merge(dfpop_idade, dfanos_idade, on=['id', 'local', 'Categoria',  'unidade', 'ano', 'Trimestre', 'AnoSedec'], how='inner')
df_idade = pd.merge(df_idade, dfforca_idade, on=['id', 'local', 'Categoria',  'unidade', 'ano', 'Trimestre', 'AnoSedec'], how='inner')
df_idade.drop(columns=['População'], inplace=True)
df_idade.to_excel("C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas Tratadas\\Tipo IDADE estadual.xlsx")

dfpop_raca.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\População RACA estadual.xlsx', index=False)
dftrab_raca.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\Trabalho RACA estadual.xlsx', index=False)
df_raca = pd.merge(dfpop_raca, dftrab_raca, on=['id', 'local', 'Categoria', 'unidade', 'ano', 'Trimestre', 'AnoSedec'], how='inner')
df_raca.to_excel("C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas Tratadas\\Tipo RAÇA estadual.xlsx")

dfpop_alfab.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\População Grau de instrução estadual.xlsx', index=False)
dftrab_alfab.to_excel('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas\\Trabalha Grau de instrução estadual.xlsx', index=False)
df_alfab = pd.merge(dfpop_alfab, dftrab_alfab, on=['id', 'local', 'Categoria', 'unidade', 'ano', 'Trimestre', 'AnoSedec'], how='inner')
df_alfab.to_excel("C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas Tratadas\\Tipo GRAU DE INSTRUÇÃO estadual.xlsx")



planilha_principal = openpyxl.Workbook()

wb_1209 = openpyxl.load_workbook('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas Tratadas\\Tipo SEXO estadual.xlsx')
wb_5918 = openpyxl.load_workbook('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas Tratadas\\Tipo IDADE estadual.xlsx')
wb_6463 = openpyxl.load_workbook('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas Tratadas\\Tipo RAÇA estadual.xlsx')
wb_6482 = openpyxl.load_workbook('C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas Tratadas\\Tipo GRAU DE INSTRUÇÃO estadual.xlsx')

aba_1209 = planilha_principal.create_sheet("SEXO")
aba_5918 = planilha_principal.create_sheet("IDADE")
aba_6463 = planilha_principal.create_sheet("RAÇA")
aba_6482 = planilha_principal.create_sheet("GRAU DE INSTRUÇÃO")


for linha in wb_1209.active.iter_rows(values_only=True):
    aba_1209.append(linha)

for linha in wb_5918.active.iter_rows(values_only=True):
    aba_5918.append(linha)
    
for linha in wb_6463.active.iter_rows(values_only=True):
    aba_6463.append(linha)
    
for linha in wb_6482.active.iter_rows(values_only=True):
    aba_6482.append(linha)
    
for aba in planilha_principal.sheetnames:
    if aba not in ["SEXO", "IDADE", "RAÇA", "GRAU DE INSTRUÇÃO"]:
        del planilha_principal[aba]
        
ajustar_bordas(planilha_principal)

lista_aba = [aba_1209, aba_5918, aba_6463, aba_6482]
for abas in lista_aba:
    ajustar_colunas(abas)
    
planilha_principal.save("C:\\Users\\LucasFreitas\\Documents\\Lucas Freitas Arquivos\\DATAHUB\\DADOS\\PNAD\\Planilhas Tratadas\\PNAD TIPOS ESTADUAL.xlsx")   

if __name__ == '__main__':
    from sql_Tipos import executar_sql 
    executar_sql()
